import json, os, gc, shutil, tempfile, traceback, zipfile, pandas as pd
from datetime import datetime
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.db.models import Count, Sum
from django.db.models.functions import Cast
from django.db.models import IntegerField
from django.contrib import messages
from django.contrib.gis.gdal import DataSource
from django.contrib.gis.geos import GEOSGeometry, MultiPolygon, Polygon
from django.contrib.gis.serializers import geojson
from django.core.serializers import serialize
from .models import *
from .forms import ShapefileUploadForm
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.db.models import Count
from openpyxl import Workbook
from .models import cde
from openpyxl.styles import Font, Border, Side
from django.contrib.auth import logout
from django.shortcuts import redirect

list_deptos = depto.objects.all()
list_munic = munic.objects.all()

@login_required
def inicio_2(request):
    if request.user.is_superuser:
        total_personal = contactoxcv.objects.count()
        total_centros = cv.objects.count()
        total_carga = cv.objects.annotate(carga_int=Cast('carga_electoral', IntegerField())).aggregate(total=Sum('carga_int'))['total'] or 0
        cantidad_cde_ejercito = cde.objects.count()
       
    else:
        perfil = PerfilUsuario.objects.get(usuario=request.user)
        id_procedencia = perfil.unidad_militar.pk
        procedencia_padre = procedencia.objects.get(id=id_procedencia)
        todos_los_ids = obtener_subprocedencias(procedencia_padre)
        todos_los_ids.add(procedencia_padre.id)  # Incluir la procedencia principal
        total_centros = cv.objects.filter(procedencia_id__in=todos_los_ids).count()
        total_carga = cv.objects.filter(procedencia__in=todos_los_ids).annotate(carga_int=Cast('carga_electoral', IntegerField())).aggregate(total=Sum('carga_int'))['total'] or 0
        cantidad_cde_ejercito = cde.objects.filter(procedencia_id__in=todos_los_ids).count()
        total_personal = contactoxcv.objects.count()

    context = {
        'CV': f"{total_centros:,}",
        'RH' : f"{total_personal:,}",
        'RUTAS' : f"{cantidad_cde_ejercito:,}",
        'ME' : f"{total_carga:,}",
    }
    return render(request, 'core/base.html', context)


@login_required
def rpt_info_gen(request):
    # Estados
    estado_data = cv.objects.values('cod_estado__descrip_corta').annotate(total=Count('id')).order_by('cod_estado__descrip_corta')
    labels_estado = [item['cod_estado__descrip_corta'] for item in estado_data]
    data_estado = [item['total'] for item in estado_data]

    # Departamentos (cantidad de CV por depto)
    depto_data_cv = (
        cv.objects
        .values('cod_depto', 'cod_depto__descrip_corta')
        .annotate(total=Count('id'))
        .order_by('cod_depto__descrip_corta')
    )
    labels_depto = [item['cod_depto__descrip_corta'] for item in depto_data_cv]
    data_depto = [item['total'] for item in depto_data_cv]
    codigos_depto = [item['cod_depto'] for item in depto_data_cv]  # 👈 Aquí

    # Carga electoral por departamento
    depto_data_carga = (
        cv.objects
        .annotate(carga_int=Cast('carga_electoral', IntegerField()))
        .values('cod_depto__descrip_corta')
        .annotate(total_carga=Sum('carga_int'))
        .order_by('cod_depto__descrip_corta')
    )
    labels_carga = [item['cod_depto__descrip_corta'] for item in depto_data_carga]
    data_carga = [item['total_carga'] for item in depto_data_carga]

    context = {
        'labels_estado': labels_estado,
        'data_estado': data_estado,
        'labels_depto': labels_depto,
        'data_depto': data_depto,
        'codigos_depto': codigos_depto,  # 👈 Incluye en contexto
        'labels_carga': labels_carga,
        'data_carga': data_carga,
    }
    return render(request, 'cv/reportes/rpt_info_gen.html', context)
@login_required
def rpt_info_situacion(request):
    estado_data = cv.objects.values('cod_estado__descrip_corta').annotate(total=Count('id')).order_by('cod_estado__descrip_corta')
    labels_estado = [item['cod_estado__descrip_corta'] for item in estado_data]
    data_estado = [item['total'] for item in estado_data]
    datos1 = (
        cde.objects
        .values('cod_estado_cde__descrip_corta')  # ajusta si `estado_cde` tiene otro campo de nombre
        .annotate(total_carga=Sum('carga_electoral'))
        .order_by('cod_estado_cde__descrip_corta')
    )
    
    labels2 = [item['cod_estado_cde__descrip_corta'] or 'Sin Estado' for item in datos1]
    datas2 = [item['total_carga'] or 0 for item in datos1]
    # Preparar los datos para el gráfico


    datos = (
        cme.objects
        .values('cod_estado_cme__descrip_corta')  # ajusta si `estado_cde` tiene otro campo de nombre
        .annotate(total_carga=Sum('carga_electoral'))
        .order_by('cod_estado_cme__descrip_corta')
    )
    
    labels1 = [item['cod_estado_cme__descrip_corta'] or 'Sin Estado' for item in datos]
    datas1 = [item['total_carga'] or 0 for item in datos]
    context = {
        'labels_estado': labels_estado,
        'data_estado': data_estado,
        'labels1': labels1,
        'datas1': datas1,
        'labels2': labels2,
        'datas2': datas2,
    }  
    return render(request, 'cv/reportes/rpt_info_situacion.html',context )
@login_required
def load_municipios(request):
    depto_id = request.GET.get('depto_id')  # Obtener la ID de la marca desde la solicitud Ajax
    municipios = munic.objects.filter(cod_depto=depto_id).order_by('descrip_corta')  # Filtrar modelos por marca
    return JsonResponse(list(municipios.values('id', 'descrip_corta')), safe=False)
@login_required
def load_cv(request):
    cod_munic = request.GET.get('cod_munic')  # Obtener la ID de la marca desde la solicitud Ajax
    centros = cv.objects.filter(cod_munic=cod_munic).order_by('nom')  # Filtrar modelos por marca
    return JsonResponse(list(centros.values('id', 'nom')), safe=False)
@login_required
def frm_princ_cv(request):
    if request.user.is_superuser:
        cvs = cv.objects.all().order_by('id')
    else:
        perfil = PerfilUsuario.objects.get(usuario=request.user)
        id_procedencia = perfil.unidad_militar.pk
        procedencia_padre = procedencia.objects.get(id=id_procedencia)
        todos_los_ids = obtener_subprocedencias(procedencia_padre)
        todos_los_ids.add(procedencia_padre.id) 
        cvs = cv.objects.filter(procedencia_id__in=todos_los_ids).order_by('id')
    context = {
        'cv' : cvs
    }
    return render(request, 'cv/form_gen_cv.html', context)
@login_required
def frm_adminxunidad_cv(request):
    if request.user.is_superuser:
        cvs = cv.objects.all().order_by('id')
    else:
        perfil = PerfilUsuario.objects.get(usuario=request.user)
        id_procedencia = perfil.unidad_militar.pk
        procedencia_padre = procedencia.objects.get(id=id_procedencia)
        todos_los_ids = obtener_subprocedencias(procedencia_padre)
        todos_los_ids.add(procedencia_padre.id) 
        cvs = cv.objects.filter(procedencia_id__in=todos_los_ids).order_by('id')
    context = {
        'cv' : cvs
    }
    return render(request, 'cv/form_admin_cv.html', context)
@login_required
def frm_mapa(request):
    return render(request, 'cv/form_mapa.html')
@login_required
def cargar_contactos(request):
    if request.method == 'POST':
        Guia_Telefonica.objects.all().delete()
        # Obtenemos el archivo subido
        archivo_excel = request.FILES['archivo_excel']
        
        # Leemos el archivo Excel con pandas
        try:
            df = pd.read_excel(archivo_excel)
            # Iteramos sobre las filas del DataFrame
            for index, row in df.iterrows():
                # Crear el contacto en la base de datos
                Guia_Telefonica.objects.create(
                    no=row['no'],
                    grado=row['grado'],
                    nom_ape=row['nom_ape'],
                    fuerza=row['fuerza'],
                    unidad=row['unidad'],
                    cargo_actual=row['cargo_actual'],
                    tel_oficina=row['tel_oficina'],
                    tel_ip=row['tel_ip'],
                    tel_celular=row['tel_celular'],
                    correo_inst=row['correo_inst'],
                    correo_personal=row['correo_personal']
                )
            return HttpResponse("Contactos cargados correctamente")
        except Exception as e:
            return HttpResponse(f"Error al cargar los contactos: {e}")
    return render(request, 'cv/cargar_contactos.html')
@login_required
def cargar_cv(request):
    if request.method == 'POST':
        cv.objects.all().delete()
        archivo_excel = request.FILES['archivo_excel']
        try:
            df = pd.read_excel(archivo_excel)
            for index, row in df.iterrows():
                try:
                    cv.objects.create(
                        cod_depto=depto.objects.get(id=row['cod_depto']),
                        cod_munic=munic.objects.get(cod_munic=row['cod_munic']),
                        cod_area=area.objects.get(id=row['cod_area']),
                        latitud=row['latitud'],
                        longitud=row['longitud'],
                        sector_electoral=row['sector_electoral'],
                        nom=row['nom'],
                        carga_electoral=str(row['carga_electoral']),
                        jrv=str(row['jrv']),
                        procedencia=procedencia.objects.get(id=row['cod_procedencia']),
                        cod_estado=estado.objects.get(id=row['cod_estado'])
                    )
                except Exception as fila_error:
                    print(f"Error en la fila {index + 2}: {fila_error}")
                    return HttpResponse(f"Error en la fila {index + 2}: {fila_error}")
            return HttpResponse("Centro de Votación cargados correctamente")
        except Exception as e:
            return HttpResponse(f"Error al procesar el archivo: {e}")
    return render(request, 'cv/cargar_contactos.html')
@login_required
def cargar_cme(request):
    if request.method == 'POST':
        cme.objects.all().delete()
        archivo_excel = request.FILES['archivo_excel']
        try:
            df = pd.read_excel(archivo_excel)
            for index, row in df.iterrows():
                try:
                    cod_munic_id = int(row['cod_munic'])
                    carga = int(float(row['carga_electoral']))
                    cme.objects.create(
                        cod_depto=depto.objects.get(id=row['depto']),
                        cod_munic=munic.objects.get(cod_munic=cod_munic_id),
                        latitud=row['latitud'],
                        longitud=row['longitud'],
                        carga_electoral=carga,
                        cod_estado_cme=estado_cme.objects.get(id=row['cod_estado'])
                    )
                except Exception as fila_error:

                    print(f"Error en la fila {index + 2}: {fila_error}")
                    return HttpResponse(f"Error en la fila {index + 2}: {fila_error}")
            return HttpResponse("Centro de Votación cargados correctamente")
        except Exception as e:
            return HttpResponse(f"Error al procesar el archivo: {e}")
    return render(request, 'cv/cargar_contactos.html')
@login_required
def guia_tel(request):
    if request.method == 'GET':
        contactos = Guia_Telefonica.objects.all()
        return render(request, 'cv/form_guia_telefonica.html', {'contactos': contactos})
@login_required
def guia_tel_cv(request):
    if request.method == 'GET':
        contactos = contactoxcv.objects.all()
        return render(request, 'cv/form_guia_telefonica_cv.html', {'contactos': contactos})
@login_required   
def registrar_manual(request):
    departamentos = depto.objects.all()
    return render(request, 'cv/form_registro.html', {'departamentos': departamentos})
@login_required
def actualizar_cv(request):
    if request.method == 'POST':
        centro_id = request.POST.get('cv')
        estado_id = request.POST.get('estado')
        nueva_img = request.FILES.get('foto_cv')
        try:
            centro = get_object_or_404(cv, id=centro_id)
            if estado_id:
                centro.cod_estado = get_object_or_404(estado, id=estado_id)
            if nueva_img:
                centro.img = nueva_img
            centro.save()
            noms = request.POST.getlist('nombre_seguridad[]')
            tels = request.POST.getlist('telefono_seguridad[]')
            img_seguridad = request.FILES.getlist('foto_seguridad[]')
            for i in range(len(noms)):
                nombre = noms[i]
                telefono = tels[i]
                foto = img_seguridad[i] if i < len(img_seguridad) else None

                contacto = contactoxcv(
                nom=nombre,
                num=telefono,
                cod_cv=centro
                )
                if foto:
                    contacto.img = foto
                contacto.save()
            
            return redirect(f"{reverse('registrar_manual')}?exito=1")
        except Exception as e:
            return JsonResponse({'error': f'Error al actualizar: {e}'}, status=500)
    return JsonResponse({'error': 'Método no permitido'}, status=405)
@login_required
def detalle_centro(request, centro_id):
    centro = cv.objects.get(id=centro_id)
    personal = contactoxcv.objects.filter(cod_cv=centro_id)

    return JsonResponse({
        "nombre": centro.nom,
        "departamento": centro.cod_depto.descrip_corta,
        "municipio": centro.cod_munic.descrip_corta,
        "sector": centro.sector_electoral,
        "carga": centro.carga_electoral,
        "unidad": centro.procedencia.descrip_larga,
        "foto_centro": centro.img.url if centro.img else "",
                "personal": [
            {
                "gradynom": p.nom,
                "telefono": p.num,
                "foto": p.img.url if p.img else ""
            }
            for p in personal
        ]
    })
@login_required
def obtener_datos_cv(request):
    cv_id = request.GET.get('cv_id')
    try:
        centro = cv.objects.select_related('cod_estado').get(id=cv_id)
        contactos = contactoxcv.objects.filter(cod_cv=centro)

        contactos_data = []
        for c in contactos:
            contactos_data.append({
                'id': c.id,
                'nom': c.nom,
                'num': c.num,
                'img_url': c.img.url if c.img else '',
            })

        data = {
            'estado_id': centro.cod_estado.id if centro.cod_estado else '',
            'img_url': centro.img.url if centro.img else '',
            'personal': contactos_data,
        }
        return JsonResponse(data)
    except cv.DoesNotExist:
        return JsonResponse({'error': 'Centro no encontrado'}, status=404)
@login_required
def eliminar_contacto(request):
    contacto_id = request.GET.get('id')
    try:
        contacto = contactoxcv.objects.get(pk=contacto_id)
        contacto.delete()
        return JsonResponse({'success': True})
    except contactoxcv.DoesNotExist:
        return JsonResponse({'error': 'Contacto no encontrado'}, status=404)
@login_required
def cargar_shapefile(request):
    if request.method == 'POST':

        form = ShapefileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            archivo_zip = request.FILES['archivo']
            try:
                # Paso 1: Extraer el ZIP en un directorio temporal
                with tempfile.TemporaryDirectory() as tmpdirname:
                    zip_path = os.path.join(tmpdirname, archivo_zip.name)
                    with open(zip_path, 'wb') as f:
                        for chunk in archivo_zip.chunks():
                            f.write(chunk)

                    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                        zip_ref.extractall(tmpdirname)

                    # Buscar el archivo .shp
                    shp_files = [f for f in os.listdir(tmpdirname) if f.endswith('.shp')]
                    if not shp_files:
                        messages.error(request, "No se encontró ningún archivo .shp.")
                        return redirect('cargar_shapefile')

                    base_name = shp_files[0][:-4]

                    # Paso 2: Copiar todos los archivos del shapefile a otra carpeta (fuera del `with`)
                    tmp_copy_dir = tempfile.mkdtemp()
                    for ext in ['.shp', '.shx', '.dbf', '.prj', '.cpg']:
                        src_file = os.path.join(tmpdirname, base_name + ext)
                        if os.path.exists(src_file):
                            shutil.copy(src_file, tmp_copy_dir)

                # Paso 3: Procesar el shapefile (fuera del bloque anterior)
                shp_path = os.path.join(tmp_copy_dir, base_name + '.shp')
                ds = DataSource(shp_path)
                layer = ds[0]
                print(layer)
                for feat in layer:
                    nombre = feat.get('shapename') or 'Sin nombre'
                    shapeid = feat.get('shapeid')
                    geom = GEOSGeometry(feat.geom.geojson)

                    if isinstance(geom, Polygon):
                        geom = MultiPolygon(geom)

                    Depto_Geo.objects.create(nombre=nombre, geom=geom, shapeid=shapeid)

                # Liberar recursos antes de eliminar archivos
                del layer
                del ds
                gc.collect()

                # Limpiar carpeta temporal manualmente
                shutil.rmtree(tmp_copy_dir, ignore_errors=True)

                messages.success(request, "Shapefile cargado correctamente.")
                return redirect('cargar_shapefile')

            except Exception as e:
                error_completo = traceback.format_exc()
                messages.error(request, f"Ocurrió un error:\n{error_completo}")
                return redirect('cargar_shapefile')
    else:
        form = ShapefileUploadForm()
    return render(request, 'cv/cargar_shapefile.html', {'form': form})
@login_required
def mapa_municipios(request):
    return render(request, 'cv/mapa_municipios.html')
@login_required
def geojson_municipios(request):
    data = serialize(
        'geojson',
        Munic_Geo.objects.all(),
        geometry_field='geom',
        fields=('nombre', 'shapeid')
    )
    return JsonResponse(data, safe=False)
@login_required
def geojson_depto(request):
    data = serialize(
        'geojson',
        Depto_Geo.objects.all(),
        geometry_field='geom',
        fields=('nombre', 'shapeid')
    )
    return JsonResponse(data, safe=False)
@login_required
def geojson_departamento(request, departamento):
    # Filtramos municipios por el nombre del departamento (en cod_munic.depto)
    municipios = Munic_Geo.objects.filter(cod_munic__cod_depto=departamento)
    features = []

    for municipio in municipios:
        # Filtro de CV relacionados al municipio actual
        cvs = cv.objects.filter(cod_munic=municipio.cod_munic)

        # Conteo por estado (según cod_estado.descripcion)
        abiertos = cvs.filter(cod_estado=3).count()
        inactivos = cvs.filter(cod_estado=1).count()
        cerrados = cvs.filter(cod_estado=4).count()
        activos = cvs.filter(cod_estado=2).count()
        cod_municipio = str(municipio.cod_munic.cod_munic)

        # Construir cada feature tipo GeoJSON
        features.append({
            "type": "Feature",
            "geometry": json.loads(municipio.geom.geojson),
            "properties": {
                "nombre": municipio.nombre,
                "abiertos": abiertos,
                "inactivos": inactivos,
                "cerrados": cerrados,
                "activos": activos,
                "codmunicipio": cod_municipio
            }
        })

    geojson_data = {
        "type": "FeatureCollection",
        "features": features
    }

    return JsonResponse(geojson_data)
@login_required  
def mapa_departamento(request):
    departamentos = depto.objects.all()
    estados_data_CDE = cde.objects.values('cod_estado_cde__descrip_corta').annotate(total=Count('id'))
    labels_CDE = [item['cod_estado_cde__descrip_corta'] for item in estados_data_CDE]
    counts_CDE = [item['total'] for item in estados_data_CDE]
    
    estados_data_CME = cme.objects.values('cod_estado_cme__descrip_corta').annotate(total=Count('id'))
    labels_CME = [item['cod_estado_cme__descrip_corta'] for item in estados_data_CME]
    counts_CME = [item['total'] for item in estados_data_CME]
    
    estados_data_CV = cv.objects.values('cod_estado__descrip_corta').annotate(total=Count('id'))
    labels_CV = [item['cod_estado__descrip_corta'] for item in estados_data_CV]
    counts_CV = [item['total'] for item in estados_data_CV]
    context = {
        'labels_CDE': labels_CDE,
        'datas_CDE': counts_CDE,
        'labels_CME': labels_CME,
        'datas_CME': counts_CME,
        'labels_CV': labels_CV,
        'datas_CV': counts_CV,
        'departamentos' : departamentos,
    }  
    return render(request, 'cv/reportes/rpt_info_depto.html', context)
@login_required
def centros_por_departamento(request, id_departamento):
    centros = cv.objects.filter(cod_munic__cod_depto=id_departamento)
    data = []
    for centro in centros:
        data.append({
            "nombre": centro.nom,
            "municipio": centro.cod_munic.descrip_corta,
            "lat": centro.latitud,
            "lng": centro.longitud,
            "estado": centro.cod_estado.descrip_corta
        })
    return JsonResponse(data, safe=False)
@login_required
def centros_por_municipio(request, cod_municipio):
    centros = cv.objects.filter(cod_munic__cod_munic=cod_municipio)
    data = []
    for centro in centros:
        contactos = contactoxcv.objects.filter(cod_cv=centro)
        contactos_data = [
            {
                'nombre': c.nom,
                'numero': c.num,
                'imagen': c.img.url if c.img else None
            }
            for c in contactos
        ]
        print(contactos_data)
        data.append({
            'id': centro.id,
            'nombre': centro.nom,
            'estado': str(centro.cod_estado),
            'contactos': contactos_data
        })
    return JsonResponse(data, safe=False)
@login_required
def obtener_personal_cv(request, centro_id):
    print("hola")
    if request.method == 'GET':
        contactos = contactoxcv.objects.filter(cod_cv=centro_id)
        personal = [

            {
                'id' : c.id,
                'nom': c.nom,
                'num': c.num,
                'img': c.img.url if c.img else None
            }
            for c in contactos
            
        ]
        print(personal)
        return JsonResponse(list(personal), safe=False)
@login_required
@csrf_exempt
@require_http_methods(["POST"])
def agregar_personal_cv(request):
    if request.method == 'POST':
        try:
            # Acceder directamente desde POST y FILES
            centro_id = request.POST.get('centro_id')
            nombre = request.POST.get('nombre')
            telefono = request.POST.get('telefono')
            foto = request.FILES.get('foto')  # Aquí sí recibes el archivo real

            centro = cv.objects.get(id=centro_id)

            nuevo = contactoxcv.objects.create(
                cod_cv=centro,
                nom=nombre,
                num=telefono,
                img=foto  # El campo `img` debe ser un ImageField en el modelo
            )

            return JsonResponse({'status': 'ok', 'id': nuevo.id})
        
        except ObjectDoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Centro no encontrado'}, status=404)
        
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)
@login_required
@csrf_exempt
def eliminar_personal(request, id):
    if request.method == 'DELETE':
        try:
            personal = contactoxcv.objects.get(id=id)
            personal.delete()
            return JsonResponse({'status': 'ok'})
        except contactoxcv.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'No encontrado'}, status=404)
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)
@login_required
def obtener_foto_cv(request, centro_id):
    try:
        centro = cv.objects.get(id=centro_id)
        url = centro.img.url if centro.img else ''
        return JsonResponse({'foto': url})
    except cv.DoesNotExist:
        return JsonResponse({'foto': ''})
@login_required
@csrf_exempt
def guardar_foto_cv(request):
    if request.method == 'POST':
        centro_id = request.POST.get('centro_id')
        nueva_foto = request.FILES.get('nueva_foto')

        try:
            centro = cv.objects.get(id=centro_id)
            centro.img = nueva_foto
            centro.save()
            return JsonResponse({'status': 'ok'})
        except cv.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'CV no encontrado'}, status=404)
@login_required        
@csrf_exempt
def guardar_ubicacion_cv(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        try:
            centro = cv.objects.get(id=data['centro_id'])
            centro.latitud = data['lat']
            centro.longitud = data['lng']
            centro.save()
            return JsonResponse({'status': 'ok'})
        except cv.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Centro no encontrado'}, status=404)
        
@csrf_exempt
def actualizar_estado_cv(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            centro = cv.objects.get(id=data['id'])
            nuevo_estado = estado.objects.get(id=data['estado'])  # o tu modelo de estado
            centro.cod_estado = nuevo_estado
            centro.save()
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)
@login_required
def detalle_departamento(request, codigo):
    # Obtener el departamento
    departamento = get_object_or_404(depto, id=codigo)
    municipios = munic.objects.filter(cod_depto=departamento).distinct()
    estados = estado.objects.all()

    labels_municipio = []
    codigos_municipio = []
    data_cv_municipio = []
    data_carga_municipio = []
    data_jrv_municipio = []

    estado_por_municipio = {e.descrip_corta: [] for e in estados}
    tabla_resumen = []

    for municipio in municipios:
        cvs = cv.objects.filter(cod_depto=departamento, cod_munic=municipio)

        if cvs.count() == 0:
            continue  # Saltar municipios sin CV

        cantidad_cv = cvs.count()
        carga_total = sum(int(c.carga_electoral) for c in cvs if c.carga_electoral and c.carga_electoral.isdigit())
        jrv_total = sum(int(c.jrv) for c in cvs if c.jrv and c.jrv.isdigit())

        labels_municipio.append(municipio.descrip_corta)
        codigos_municipio.append(municipio.cod_munic)
        data_cv_municipio.append(cantidad_cv)
        data_carga_municipio.append(carga_total)
        data_jrv_municipio.append(jrv_total)

        # Agregar a la tabla resumen
        tabla_resumen.append({
            'municipio': municipio.descrip_corta,
            'cv': cantidad_cv,
            'jrv': jrv_total,
            'carga': carga_total
        })

        # Conteo por estado
        for e in estados:
            cantidad_estado = cvs.filter(cod_estado=e).count()
            estado_por_municipio[e.descrip_corta].append(cantidad_estado)

    context = {
        'departamento': departamento,
        'labels_municipio': labels_municipio,
        'codigos_municipio': codigos_municipio,
        'data_cv_municipio': data_cv_municipio,
        'data_carga_municipio': data_carga_municipio,
        'data_jrv_municipio': data_jrv_municipio,
        'estado_nombres': list(estado_por_municipio.keys()),
        'estado_datasets': list(estado_por_municipio.values()),
        'tabla_resumen': tabla_resumen,
    }
    return render(request, 'cv/reportes/detalle_departamento.html', context)
@login_required
def frm_admin_cde(request):
    cdes = cde.objects.all().order_by('id')
    datos = (
        cde.objects
        .values('cod_estado_cde__descrip_corta')  # ajusta si `estado_cde` tiene otro campo de nombre
        .annotate(total_carga=Sum('carga_electoral'))
        .order_by('cod_estado_cde__descrip_corta')
    )
    labels1 = [item['cod_estado_cde__descrip_corta'] or 'Sin Estado' for item in datos]
    datas1 = [item['total_carga'] or 0 for item in datos]
    
    estados_data = cde.objects.values('cod_estado_cde__descrip_corta').annotate(total=Count('id'))
    labels = [item['cod_estado_cde__descrip_corta'] for item in estados_data]
    counts = [item['total'] for item in estados_data]
    context = {
        'cde' : cdes,
        'departamentos': depto.objects.all(),
        'municipios': munic.objects.all(),
        'procedencias': procedencia.objects.all(),
        'estados': estado_cde.objects.all(),
        'labels': labels,
        'counts': counts,
        'labels1': labels1,
        'datas1': datas1,
    }
    return render(request, 'cde/form_admin_cde.html', context)
@login_required
def registrar_cde(request):
    try:
        id_cde = request.POST.get('id_cde')  # Viene oculto del formulario si es edición
        if id_cde:
            cdes = cde.objects.get(id=id_cde)
        else:
            cdes = cde()
        cdes.cod_depto_id = request.POST.get('departamento')
        cdes.cod_munic_id = request.POST.get('municipio')
        cdes.procedencia_id = request.POST.get('procedencia')
        cdes.carga_electoral = request.POST.get('carga_electoral')
        cdes.hora_desp_salida = request.POST.get('hora_desp_salida') or None
        cdes.hora_desp_llegada = request.POST.get('hora_desp_llegada') or None
        cdes.hora_redesp_salida = request.POST.get('hora_redesp_salida') or None
        cdes.hora_redesp_llegada = request.POST.get('hora_redesp_llegada') or None
        cdes.obs = request.POST.get('obs') or ""
        if 'img' in request.FILES:
            cdes.img = request.FILES['img']
        cdes.save()
        mensaje = 'Registro actualizado correctamente.' if id_cde else 'Registro guardado correctamente.'
        return JsonResponse({'success': True, 'message': mensaje})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
@login_required
def eliminar_cde(request, pk):
    cdes = get_object_or_404(cde, pk=pk)
    cdes.delete()
    messages.success(request, "Consejo Departamental eliminado exitosamente.")
    return redirect(frm_admin_cde)  # Reemplaza por el nombre real
@login_required
def obtener_cde(request, id):
    print(id)
    cdes = cde.objects.get(pk=id)
    data = {
        'id': cdes.id,
        'cod_depto': cdes.cod_depto.id if cdes.cod_depto else None,
        'cod_munic': cdes.cod_munic.id if cdes.cod_munic else None,
        'procedencia': cdes.procedencia.id if cdes.procedencia else None,
        'carga_electoral': cdes.carga_electoral,
        'hora_desp_salida': cdes.hora_desp_salida.strftime('%Y-%m-%dT%H:%M') if cdes.hora_desp_salida else '',
        'hora_desp_llegada': cdes.hora_desp_llegada.strftime('%Y-%m-%dT%H:%M') if cdes.hora_desp_llegada else '',
        'hora_redesp_salida': cdes.hora_redesp_salida.strftime('%Y-%m-%dT%H:%M') if cdes.hora_redesp_salida else '',
        'hora_redesp_llegada': cdes.hora_redesp_llegada.strftime('%Y-%m-%dT%H:%M') if cdes.hora_redesp_llegada else '',
        'latitud': cdes.latitud,
        'longitud': cdes.longitud,
        'obs': cdes.obs,
    }
    return JsonResponse(data)
@login_required
def obtener_foto_cde(request, cde_id):
    try:
        cdes = cde.objects.get(id=cde_id)
        url = cdes.img.url if cdes.img else ''
        return JsonResponse({'foto': url})
    except cv.DoesNotExist:
        return JsonResponse({'foto': ''})
@login_required   
@csrf_exempt
def guardar_foto_cde(request):
    if request.method == 'POST':
        centro_id = request.POST.get('centro_id')
        nueva_foto = request.FILES.get('nueva_foto')

        try:
            centro = cde.objects.get(id=centro_id)
            centro.img = nueva_foto
            centro.save()
            return JsonResponse({'status': 'ok'})
        except cv.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'CV no encontrado'}, status=404)
@login_required
@csrf_exempt
def guardar_ubicacion_cde(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        try:
            centro = cde.objects.get(id=data['centro_id'])
            centro.latitud = data['lat']
            centro.longitud = data['lng']
            centro.save()
            return JsonResponse({'status': 'ok'})
        except cv.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Centro no encontrado'}, status=404)
        
@login_required
def obtener_personal_cde(request, centro_id):
    if request.method == 'GET':
        contactos = contactoxcv.objects.filter(cod_cde=centro_id)
        personal = [

            {
                'id' : c.id,
                'nom': c.nom,
                'num': c.num,
                'img': c.img.url if c.img else None
            }
            for c in contactos
            
        ]
        print(personal)
        return JsonResponse(list(personal), safe=False)

@csrf_exempt
@require_http_methods(["POST"])
def agregar_personal_cde(request):
    if request.method == 'POST':
        try:
            # Acceder directamente desde POST y FILES
            centro_id = request.POST.get('centro_id')
            nombre = request.POST.get('nombre')
            telefono = request.POST.get('telefono')
            foto = request.FILES.get('foto')  # Aquí sí recibes el archivo real

            centro = cde.objects.get(id=centro_id)

            nuevo = contactoxcv.objects.create(
                cod_cde=centro,
                nom=nombre,
                num=telefono,
                img=foto  # El campo `img` debe ser un ImageField en el modelo
            )

            return JsonResponse({'status': 'ok', 'id': nuevo.id})
        
        except ObjectDoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Centro no encontrado'}, status=404)
        
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

@csrf_exempt
def actualizar_estado_cde(request):
    if request.method == 'POST':
        try:
            
            data = json.loads(request.body)
            print(data['id'])
            print(data['estado'])
            centro = cde.objects.get(id=data['id'])
            nuevo_estado = estado_cde.objects.get(id=data['estado'])  # o tu modelo de estado
            centro.cod_estado_cde = nuevo_estado
            centro.save()
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)
@login_required
def exportar_cde_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "CDE"

    # Encabezados
    columnas = [
        "No", "Departamento","Municipio","Estado CDE",  "Carga Electoral",
        "Hora Desp. Salida", "Hora Desp. Llegada",
        "Hora Redesp. Salida", "Hora Redesp. Llegada",
          "Procedencia"
    ]

    ws.append(columnas)

    # Estilos
    borde_fino = Side(border_style="thin", color="000000")
    borde_grueso = Side(border_style="medium", color="000000")

    # Aplicar estilos a encabezados
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = Font(bold=True)
            cell.border = Border(
                top=borde_grueso, left=borde_grueso,
                right=borde_grueso, bottom=borde_grueso
            )

    # Datos
    for i, obj in enumerate(cde.objects.all(), start=1):
        fila = [
            i,
            obj.cod_depto.descrip_corta if obj.cod_depto else '',
            obj.cod_munic.descrip_corta if obj.cod_munic else '',
            obj.cod_estado_cde.descrip_corta if obj.cod_estado_cde else '',
            obj.carga_electoral,
            obj.hora_desp_salida_formateada(),
            obj.hora_desp_llegada_formateada(),
            obj.hora_redesp_salida_formateada(),
            obj.hora_redesp_llegada_formateada(),
            obj.procedencia.descrip_corta if obj.procedencia else '',
        ]
        ws.append(fila)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(
                top=borde_fino, left=borde_fino,
                right=borde_fino, bottom=borde_fino
            )
    # Respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=CDE_exportado.xlsx'
    wb.save(response)
    return response

def cerrar_sesion(request):
    logout(request)
    return redirect(inicio_2) 

def obtener_subprocedencias(procedencia):
    subprocedencias = set()
    for sub in procedencia.subcategorias.all():
        subprocedencias.add(sub.id)
        subprocedencias.update(obtener_subprocedencias(sub))
    return subprocedencias

def frm_admin_cme(request):
    cdes = cme.objects.all().order_by('id')
    datos = (
        cme.objects
        .values('cod_estado_cme__descrip_corta')  # ajusta si `estado_cde` tiene otro campo de nombre
        .annotate(total_carga=Sum('carga_electoral'))
        .order_by('cod_estado_cme__descrip_corta')
    )
    labels1 = [item['cod_estado_cme__descrip_corta'] or 'Sin Estado' for item in datos]
    datas1 = [item['total_carga'] or 0 for item in datos]
    # Preparar los datos para el gráfico
    estados_data = cme.objects.values('cod_estado_cme__descrip_corta').annotate(total=Count('id'))
    labels = [item['cod_estado_cme__descrip_corta'] for item in estados_data]
    counts = [item['total'] for item in estados_data]
    context = {
        'cme' : cdes,
        'departamentos': depto.objects.all(),
        'municipios': munic.objects.all(),
        'procedencias': procedencia.objects.all(),
        'estados': estado_cde.objects.all(),
        'labels': labels,
        'counts': counts,
        'labels1': labels1,
        'datas1': datas1,
    }
    return render(request, 'cme/form_admin_cme.html', context)
@login_required
def registrar_cme(request):
    try:
        id_cme = request.POST.get('id_cme')  # Viene oculto del formulario si es edición
        if id_cme:
            cdes = cme.objects.get(id=id_cme)
        else:
            cdes = cme()
            cdes.longitud = -86.715087890625
            cdes.latitud = 13.854080569524527
              
        cdes.cod_depto_id = request.POST.get('departamento')
        cdes.cod_munic_id = request.POST.get('municipio')
        cdes.procedencia_id = request.POST.get('procedencia')
        cdes.carga_electoral = request.POST.get('carga_electoral')
        cdes.hora_desp_salida = request.POST.get('hora_desp_salida') or None
        cdes.hora_desp_llegada = request.POST.get('hora_desp_llegada') or None
        cdes.hora_redesp_salida = request.POST.get('hora_redesp_salida') or None
        cdes.hora_redesp_llegada = request.POST.get('hora_redesp_llegada') or None
        cdes.obs = request.POST.get('obs') or ""
        if 'img' in request.FILES:
            cdes.img = request.FILES['img']
        cdes.save()
        mensaje = 'Registro actualizado correctamente.' if id_cme else 'Registro guardado correctamente.'
        return JsonResponse({'success': True, 'message': mensaje})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
@login_required
def eliminar_cme(request, pk):
    cdes = get_object_or_404(cme, pk=pk)
    cdes.delete()
    messages.success(request, "Consejo Municipal eliminado exitosamente.")
    return redirect(frm_admin_cme)  # Reemplaza por el nombre real
@login_required
def obtener_cme(request, id):
    cdes = cme.objects.get(pk=id)
    data = {
        'id': cdes.id,
        'cod_depto': cdes.cod_depto.id if cdes.cod_depto else None,
        'cod_munic': cdes.cod_munic.id if cdes.cod_munic else None,
        'procedencia': cdes.procedencia.id if cdes.procedencia else None,
        'carga_electoral': cdes.carga_electoral,
        'hora_desp_salida': cdes.hora_desp_salida.strftime('%Y-%m-%dT%H:%M') if cdes.hora_desp_salida else '',
        'hora_desp_llegada': cdes.hora_desp_llegada.strftime('%Y-%m-%dT%H:%M') if cdes.hora_desp_llegada else '',
        'hora_redesp_salida': cdes.hora_redesp_salida.strftime('%Y-%m-%dT%H:%M') if cdes.hora_redesp_salida else '',
        'hora_redesp_llegada': cdes.hora_redesp_llegada.strftime('%Y-%m-%dT%H:%M') if cdes.hora_redesp_llegada else '',
        'latitud': cdes.latitud,
        'longitud': cdes.longitud,
    }
    return JsonResponse(data)
@login_required
def obtener_foto_cme(request, cme_id):
    try:
        cdes = cme.objects.get(id=cme_id)
        url = cdes.img.url if cdes.img else ''
        return JsonResponse({'foto': url})
    except cv.DoesNotExist:
        return JsonResponse({'foto': ''})
@login_required   
@csrf_exempt
def guardar_foto_cme(request):
    if request.method == 'POST':
        centro_id = request.POST.get('centro_id')
        nueva_foto = request.FILES.get('nueva_foto')

        try:
            centro = cme.objects.get(id=centro_id)
            centro.img = nueva_foto
            centro.save()
            return JsonResponse({'status': 'ok'})
        except cv.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'CV no encontrado'}, status=404)
@login_required
@csrf_exempt
def guardar_ubicacion_cme(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        try:
            centro = cme.objects.get(id=data['centro_id'])
            centro.latitud = data['lat']
            centro.longitud = data['lng']
            centro.save()
            return JsonResponse({'status': 'ok'})
        except cv.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Centro no encontrado'}, status=404)
        
@login_required
def obtener_personal_cme(request, centro_id):
    if request.method == 'GET':
        contactos = contactoxcv.objects.filter(cod_cme=centro_id)
        personal = [

            {
                'id' : c.id,
                'nom': c.nom,
                'num': c.num,
                'img': c.img.url if c.img else None
            }
            for c in contactos
            
        ]
        print(personal)
        return JsonResponse(list(personal), safe=False)

@csrf_exempt
@require_http_methods(["POST"])
def agregar_personal_cme(request):
    if request.method == 'POST':
        try:
            # Acceder directamente desde POST y FILES
            centro_id = request.POST.get('centro_id')
            nombre = request.POST.get('nombre')
            telefono = request.POST.get('telefono')
            foto = request.FILES.get('foto')  # Aquí sí recibes el archivo real

            centro = cme.objects.get(id=centro_id)

            nuevo = contactoxcv.objects.create(
                cod_cme=centro,
                nom=nombre,
                num=telefono,
                img=foto  # El campo `img` debe ser un ImageField en el modelo
            )

            return JsonResponse({'status': 'ok', 'id': nuevo.id})
        
        except ObjectDoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Centro no encontrado'}, status=404)
        
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

@csrf_exempt
def actualizar_estado_cme(request):
    if request.method == 'POST':
        try:
            
            data = json.loads(request.body)
            centro = cme.objects.get(id=data['id'])
            nuevo_estado = estado_cme.objects.get(id=data['estado'])  # o tu modelo de estado
            centro.cod_estado_cme = nuevo_estado
            centro.save()
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)
@login_required
def exportar_cme_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "CDE"

    # Encabezados
    columnas = [
        "No", "Departamento","Municipio","Estado CME",  "Carga Electoral",
        "Hora Desp. Salida", "Hora Desp. Llegada",
        "Hora Redesp. Salida", "Hora Redesp. Llegada",
          "Procedencia"
    ]

    ws.append(columnas)

    # Estilos
    borde_fino = Side(border_style="thin", color="000000")
    borde_grueso = Side(border_style="medium", color="000000")

    # Aplicar estilos a encabezados
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = Font(bold=True)
            cell.border = Border(
                top=borde_grueso, left=borde_grueso,
                right=borde_grueso, bottom=borde_grueso
            )

    # Datos
    for i, obj in enumerate(cme.objects.all(), start=1):
        fila = [
            i,
            obj.cod_depto.descrip_corta if obj.cod_depto else '',
            obj.cod_munic.descrip_corta if obj.cod_munic else '',
            obj.cod_estado_cme.descrip_corta if obj.cod_estado_cme else '',
            obj.carga_electoral,
            obj.hora_desp_salida_formateada(),
            obj.hora_desp_llegada_formateada(),
            obj.hora_redesp_salida_formateada(),
            obj.hora_redesp_llegada_formateada(),
            obj.procedencia.descrip_corta if obj.procedencia else '',
        ]
        ws.append(fila)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(
                top=borde_fino, left=borde_fino,
                right=borde_fino, bottom=borde_fino
            )
    # Respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=cme_exportado.xlsx'
    wb.save(response)
    return response

def exportar_cve_excel(request):
    if request.user.is_superuser:
        cvs = cv.objects.all().order_by('id')
    else:
        perfil = PerfilUsuario.objects.get(usuario=request.user)
        id_procedencia = perfil.unidad_militar.pk
        procedencia_padre = procedencia.objects.get(id=id_procedencia)
        todos_los_ids = obtener_subprocedencias(procedencia_padre)
        todos_los_ids.add(procedencia_padre.id) 
        cvs = cv.objects.filter(procedencia_id__in=todos_los_ids).order_by('id')
    wb = Workbook()
    ws = wb.active
    ws.title = "CV"
    columnas = [
        "No", "Departamento","Municipio","Sector", "Nombre", "Carga Electoral","Latitud", "Longitud", "Estado", "Unidad"]
    ws.append(columnas)
    borde_fino = Side(border_style="thin", color="000000")
    borde_grueso = Side(border_style="medium", color="000000")
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = Font(bold=True)
            cell.border = Border(
                top=borde_grueso, left=borde_grueso,
                right=borde_grueso, bottom=borde_grueso
            )
    for i, obj in enumerate(cvs,  start=1):
        fila = [
            i,
            obj.cod_depto.descrip_corta if obj.cod_depto else '',
            obj.cod_munic.descrip_corta if obj.cod_munic else '',
            obj.sector_electoral,
            obj.nom,
            obj.carga_electoral,
            obj.latitud,
            obj.longitud,
            obj.cod_estado.descrip_corta if obj.cod_estado else '',
            obj.procedencia.descrip_corta if obj.procedencia else ''
        ]
        ws.append(fila)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(
                top=borde_fino, left=borde_fino,
                right=borde_fino, bottom=borde_fino
            )
    # Respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=cve_exportado.xlsx'
    wb.save(response)
    return response


def api_cv_tabla(request):
    # Igual que antes, con permisos y filtros
    if request.user.is_superuser:
        cvs = cv.objects.all().order_by('id')
    else:
        perfil = PerfilUsuario.objects.get(usuario=request.user)
        id_procedencia = perfil.unidad_militar.pk
        procedencia_padre = procedencia.objects.get(id=id_procedencia)
        todos_los_ids = obtener_subprocedencias(procedencia_padre)
        todos_los_ids.add(procedencia_padre.id) 
        cvs = cv.objects.filter(procedencia_id__in=todos_los_ids).order_by('id')

    datos_tabla = []
    for obj in cvs:
        datos_tabla.append({
            'id': obj.id,
            'cod_depto': obj.cod_depto.descrip_corta if obj.cod_depto else '',
            'cod_munic': obj.cod_munic.descrip_corta if obj.cod_munic else '',
            'sector_electoral': obj.sector_electoral,
            'nom': obj.nom,
            'carga_electoral': obj.carga_electoral,
            'cod_estado': str(obj.cod_estado),
            'procedencia': obj.procedencia.descrip_corta if obj.procedencia else '',
            'latitud': obj.latitud,
            'longitud': obj.longitud,
        })

    return JsonResponse({'data': datos_tabla})


def api_cv_mapa(request):
    # Mismo filtro
    if request.user.is_superuser:
        cvs = cv.objects.all().order_by('id')
    else:
        perfil = PerfilUsuario.objects.get(usuario=request.user)
        id_procedencia = perfil.unidad_militar.pk
        procedencia_padre = procedencia.objects.get(id=id_procedencia)
        todos_los_ids = obtener_subprocedencias(procedencia_padre)
        todos_los_ids.add(procedencia_padre.id) 
        cvs = cv.objects.filter(procedencia_id__in=todos_los_ids).order_by('id')

    puntos = []
    for obj in cvs:
        puntos.append({
            'id': obj.id,
            'latitud': obj.latitud,
            'longitud': obj.longitud,
            'nombre': obj.nom,
            # añade más datos para popup si quieres
        })

    return JsonResponse({'data': puntos})

